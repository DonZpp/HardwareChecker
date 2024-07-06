import openpyxl
import GetConfig


WORK_DIR = GetConfig.WORK_DIR
CONFIG_FILE = GetConfig.CONFIG_FILE
CONFIG_SHEET = GetConfig.CONFIG_SHEET

NUM_COL_TYPE = GetConfig.NUM_COL_TYPE
NUM_COL_KEY = GetConfig.NUM_COL_KEY
NUM_COL_VAL = GetConfig.NUM_COL_VAL
LIST_ITEM_NUM = GetConfig.NUM_COL_LIST_ITEM_NUM
ITEM_RECORDS_NUM = GetConfig.NUM_COL_LIST_ITEM_RECORDS_NUM

TYPE_LIST = GetConfig.TYPE_LIST
TYPE_LIST_ITEM = GetConfig.TYPE_LIST_ITEM
TYPE_RECORDS = GetConfig.TYPE_RECORDS


# 直接raise和捕获Exception会导致很多原生的Exception被处理，
# 因此声明一个新的LocalException来避免
class LocalException(Exception):
    def __init__(self, *args: object) -> None:
        super().__init__(*args)


# 读取列表检查项的某一个检查点的参数
def ReadListItem(lst : list, ws : openpyxl.worksheet.worksheet, row, nItemRec):
    i = 0
    dic = {}
    while i < nItemRec:
        strKey = ws.cell(row = row, column = NUM_COL_KEY).value
        strVal = ws.cell(row = row, column = NUM_COL_VAL).value
        dic[strKey] = strVal
        i = i + 1
        row = row + 1
    lst.append(dic)
    return row


# 读取列表检查项（有多个相同的点需要检查）
def ReadList(dicConf : dict, ws : openpyxl.worksheet.worksheet, row):
    strListName = ws.cell(row = row, column = NUM_COL_KEY).value
    dicConf[strListName] = list()
    nItemNum = int(ws.cell(row = row, column = LIST_ITEM_NUM).value)
    nItemRec = int(ws.cell(row = row, column = ITEM_RECORDS_NUM).value)
    i = 0
    row = row + 1
    while i < nItemNum:
        row = ReadListItem(dicConf[strListName], ws, row, nItemRec)
        i = i + 1
    return row


# 读取某一个检查项，该检查项只有一个数据需要检查
def ReadRecords(dicConf : dict, ws : openpyxl.worksheet.worksheet, row):
    strKey = ws.cell(row=row, column=NUM_COL_KEY).value
    strVal = ws.cell(row=row, column=NUM_COL_VAL).value
    dicConf[strKey] = strVal
    return row + 1


ReadObjFunc = {
    TYPE_LIST : ReadList,
    TYPE_RECORDS : ReadRecords
}


# 读取Excel配置表
def ReadConfigExcel():
    wb = openpyxl.open(WORK_DIR + CONFIG_FILE)
    ws = wb[CONFIG_SHEET]
    i = 1

    dicConf = {}
    while ws.cell(row = i, column=1).value != None:
        strType = ws.cell(row = i, column=NUM_COL_TYPE).value
        i = ReadObjFunc[strType](dicConf, ws, i)
    return dicConf


# 比对列表检查项的某一项
def CheckListItem(dicItemLocal, dicItemModel):
    for k, v in dicItemLocal.items():
        if type(v) != type(list()):
            if v != dicItemModel[k]:
                return False
    return True


# 比对列表检查项
def CheckList(lstLocal, lstModel : list, strKey):
    # lstModel，不影响原lstModel
    lstBakModel = list()
    for itemModel in lstModel:
        lstBakModel.append(itemModel)
    setEq = list()
    for itemLocal in lstLocal:
        for itemModel in lstBakModel:
            if CheckListItem(itemLocal, itemModel) :
                setEq.append(itemLocal)
                # 检查完一项就移除一项，否则如果lstLocal中的多项与lstBakModel中的一项相等，会出现反复比对的bug
                lstModel.remove(itemModel)
                break
    if len(setEq) != len(lstLocal):
        raise LocalException('sth error!'+ strKey+ '     Local:'+ lstLocal+ '     Model:'+ lstBakModel)


# 检查入口
def Check():
    dicModel = ReadConfigExcel()
    dicLocal = GetConfig.GetConfig()

    for strCheckItemName, checkItemConf in dicLocal.items() :
        if strCheckItemName in dicModel.keys():
            # TODO 该检查封装成函数以应对变化
            if type(checkItemConf) != type(list()):
                if checkItemConf != dicModel[strCheckItemName]:
                    strErr = 'sth error!'+ strCheckItemName+ '  Local:'+ dicLocal[strCheckItemName]+ '    Model:'+ dicModel[strCheckItemName]
                    raise LocalException(strErr)
            else:
                if(type(dicModel[strCheckItemName]) != type(list())):
                    raise LocalException('sth error!' +  strCheckItemName+ '   Local:'+ dicLocal[strCheckItemName]+ '     Model:'+ dicModel[strCheckItemName])
                else:
                    CheckList(checkItemConf, dicModel[strCheckItemName], strCheckItemName)


try:
    Check()
    print("Done!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
except LocalException as e:
    print(e.args)