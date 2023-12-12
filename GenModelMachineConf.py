import GetConfig
import openpyxl


dicConf = GetConfig.GetConfig()
WORK_DIR = GetConfig.WORK_DIR
CONFIG_FILE = GetConfig.CONFIG_FILE
CONFIG_SHEET = GetConfig.CONFIG_SHEET


NUM_COL_TYPE = GetConfig.NUM_COL_TYPE
NUM_COL_KEY = GetConfig.NUM_COL_KEY
NUM_COL_VAL = GetConfig.NUM_COL_VAL
LIST_ITEM_NUM = GetConfig.NUM_COL_LIST_ITEM_NUM
ITEM_RECORDS_NUM = GetConfig.NUM_COL_LIST_ITEM_RECORDS_NUM


TYPE_LIST = GetConfig.TYPE_LIST
TYPE_RECORDS = GetConfig.TYPE_RECORDS
TYPE_LIST_ITEM = GetConfig.TYPE_LIST_ITEM

#SEPARATE_CHAR = ','


#def StrTransfer(strOrg : str):
#    return strOrg.replace(SEPARATE_CHAR, SEPARATE_CHAR + SEPARATE_CHAR)


# you must run this function in a model machine
def WriteToFile():
    wb = openpyxl.Workbook()
    wb.create_sheet(CONFIG_SHEET)
    ws = wb[CONFIG_SHEET]
    i = 1
    for key, value in dicConf.items():
        if type(value) == type(list()):
            ws.cell(row = i, column=NUM_COL_TYPE).value = TYPE_LIST
            ws.cell(row = i, column = NUM_COL_KEY).value = key
            ws.cell(row = i, column = LIST_ITEM_NUM).value = str(len(value))
            ws.cell(row = i, column = ITEM_RECORDS_NUM).value = str(len(value[0]))
            for  lstVal in value:
                for key2, value2 in lstVal.items():
                    i = i + 1
                    ws.cell(row = i, column = NUM_COL_TYPE).value = TYPE_LIST_ITEM
                    ws.cell(row = i, column=NUM_COL_KEY).value = str(key2)
                    ws.cell(row = i, column = NUM_COL_VAL).value = str(value2)
        else:
            ws.cell(row = i, column = NUM_COL_TYPE).value = TYPE_RECORDS
            ws.cell(row = i, column = NUM_COL_KEY).value = str(key)
            ws.cell(row = i, column = NUM_COL_VAL).value = str(value)
        i = i + 1
    wb.save(WORK_DIR + CONFIG_FILE)


WriteToFile()